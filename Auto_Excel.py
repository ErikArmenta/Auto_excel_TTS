# -*- coding: utf-8 -*-
"""
Created on Wed Aug 12 15:57:42 2026

@author: acer
"""

import streamlit as st
import pandas as pd
import openpyxl
import datetime
import re
import io
from supabase import create_client, Client

# --- CONFIGURACIÓN DE PÁGINA ---
st.set_page_config(page_title="Tecma - MTC Automation Desk", layout="wide", page_icon="⚙️")
st.title("⚙️ MTC Southbound - Automation Desk")

# --- CONEXIÓN A SUPABASE ---
@st.cache_resource
def init_supabase() -> Client:
    url = st.secrets["SUPABASE_URL"]
    key = st.secrets["SUPABASE_KEY"]
    return create_client(url, key)

supabase = init_supabase()

# --- FUNCIONES CORE DE LIMPIEZA ---
def clean_val(val):
    if pd.isna(val): return None
    if isinstance(val, str) and val.strip().lower() in ['nd', 'n/a', '#n/a', 'na', '-', '']: return None
    if isinstance(val, pd.Timestamp): return val.to_pydatetime()
    return val

def clean_date(val):
    if pd.isna(val): return None
    if isinstance(val, str) and val.strip().lower() in ['nd', 'n/a', '#n/a', 'na', '-', '']: return None
    try:
        dt = pd.to_datetime(val)
        return dt.date() if not pd.isna(dt) else None
    except: return None

# --- CREACIÓN DE MÓDULOS (TABS) ---
tab1, tab2, tab3 = st.tabs(["1️⃣ Rastreo de Repetidos", "2️⃣ Cruce FECHAS GENERAL", "3️⃣ Validador y Carga a BD"])

# ==========================================
# MÓDULO 1: CONTENEDORES REPETIDOS
# ==========================================
with tab1:
    st.header("1️⃣ Rastreo de Contenedores Repetidos")
    st.markdown("Sube múltiples reportes semanales para detectar si un contenedor ha sido facturado varias veces.")
    
    archivos_mod1 = st.file_uploader("Sube reportes (Week 26, 27, etc.)", type=['xlsx'], accept_multiple_files=True, key="mod1")
    
    if archivos_mod1 and st.button("Generar Reporte de Repetidos", type="primary"):
        all_data = []
        file_lists = {}
        
        with st.spinner("Analizando archivos..."):
            for f in archivos_mod1:
                df = pd.read_excel(f, header=4)
                container_col = next((col for col in df.columns if 'CONTAINER' in str(col).upper()), None)
                
                if container_col:
                    containers = df[container_col].dropna().astype(str).str.strip()
                    containers = containers[containers.str.len() == 11] # Filtrar basura, solo 11 chars
                    
                    file_lists[f.name] = containers.tolist()
                    
                    # Extraer semana
                    match = re.search(r'(wk-|week\s|wk)(\d{2})', f.name, re.IGNORECASE)
                    short_name = f"WK{match.group(2)}" if match else f.name[:10]
                    
                    for c in containers:
                        all_data.append({'Container': c, 'File': short_name})
            
            if all_data:
                df_all = pd.DataFrame(all_data)
                summary = pd.crosstab(df_all['Container'], df_all['File'])
                summary['Total Repeticiones'] = summary.sum(axis=1)
                
                file_cols = [c for c in summary.columns if c != 'Total Repeticiones']
                summary['Archivos donde aparece'] = summary.apply(lambda row: ", ".join([f"{c} (x{row[c]})" if row[c] > 1 else c for c in file_cols if row[c] > 0]), axis=1)
                
                summary = summary.sort_values(by='Total Repeticiones', ascending=False).reset_index()
                duplicates = summary[summary['Total Repeticiones'] > 1]
                
                # Guardar en memoria (BytesIO) para descarga
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    duplicates.to_excel(writer, sheet_name='Repetidos', index=False)
                    summary.to_excel(writer, sheet_name='Todos los Contenedores', index=False)
                    for fname, clist in file_lists.items():
                        pd.DataFrame({'CONTAINER': clist}).to_excel(writer, sheet_name=fname[:31], index=False)
                
                st.success(f"¡Análisis completado! Se encontraron {len(duplicates)} contenedores repetidos.")
                st.download_button(label="📥 Descargar Reporte (Excel)", data=output.getvalue(), file_name="Reporte_Repetidos.xlsx", mime="application/vnd.ms-excel")

# ==========================================
# MÓDULO 2: CRUCE FECHAS GENERAL
# ==========================================
with tab2:
    st.header("2️⃣ Cruce con Base AMTC 7")
    st.markdown("Transfiere fechas de entrega, carga y retorno desde la Base de Datos AMTC al formato General.")
    
    col1, col2 = st.columns(2)
    with col1:
        base_mod2 = st.file_uploader("Sube 'Nueva base de datos AMTC (7)'", type=['xlsx'], key="mod2_base")
    with col2:
        target_mod2 = st.file_uploader("Sube 'FECHAS CONTRA GENERAL'", type=['xlsx'], key="mod2_target")
        
    if base_mod2 and target_mod2 and st.button("Ejecutar Cruce de Fechas", type="primary"):
        with st.spinner("Procesando y cruzando información..."):
            df_base = pd.read_excel(base_mod2, sheet_name='TJ-LB')
            
            base_dict = {}
            for index, row in df_base.iterrows():
                container = str(row.iloc[0]).strip()
                if container == 'nan' or not container: continue
                base_dict[container] = {
                    'amtc_delivery': clean_val(row.get('AMTC Delivery (Loaded)')),
                    'amtc_unload': clean_val(row.get('AMTC Unload Notification (Empty)')),
                    'empty_return': clean_val(row.get('Emty Return')) 
                }
                
            wb = openpyxl.load_workbook(target_mod2)
            ws = wb.active
            
            headers = {str(cell.value).strip().upper().replace('\n', ' '): col_idx for col_idx, cell in enumerate(ws[1], start=1) if cell.value}
            
            col_container = next((idx for h, idx in headers.items() if 'CONTAINER' in h), None)
            col_mtc_delivery = next((idx for h, idx in headers.items() if 'MTC DELIVERY' in h), None)
            col_empty_ready = next((idx for h, idx in headers.items() if 'EMPTY READY' in h), None)
            col_g = 7 
            ws.cell(row=1, column=col_g).value = 'Emty Return (BD)'
            
            for row in range(2, ws.max_row + 1):
                if not col_container: break
                container_cell = ws.cell(row=row, column=col_container)
                container_val = str(container_cell.value).strip() if container_cell.value else None
                
                if container_val and container_val in base_dict:
                    base_data = base_dict[container_val]
                    if col_mtc_delivery: ws.cell(row=row, column=col_mtc_delivery).value = base_data['amtc_delivery']
                    if col_empty_ready: ws.cell(row=row, column=col_empty_ready).value = base_data['amtc_unload']
                    ws.cell(row=row, column=col_g).value = base_data['empty_return']
            
            output = io.BytesIO()
            wb.save(output)
            st.success("Cruce exitoso. Archivo listo para descarga.")
            st.download_button(label="📥 Descargar FECHAS ACTUALIZADO", data=output.getvalue(), file_name="FECHAS_CONTRA_GENERAL_Actualizado.xlsx", mime="application/vnd.ms-excel")

# ==========================================
# MÓDULO 3: VALIDADOR Y BASE DE DATOS
# ==========================================
with tab3:
    st.header("3️⃣ Validador AMTC 6 y Carga a Supabase")
    st.markdown("Valida fechas contra la base AMTC 6 y sube los reportes limpios a la Base de Datos PostgreSQL.")
    
    # 1. Input de la llave relacional y consulta automática
    input_semana = st.text_input("🔑 Identificador de Semana (Ej. WK-26):", "WK-26").strip().upper()
    
    # Consulta silenciosa a Supabase usando count='exact' para no descargar datos pesados, solo el número de filas
    res = supabase.table('mtc_southbound').select('CONTAINER', count='exact').eq('SEMANA', input_semana).execute()
    registros_existentes = res.count
    
    puede_subir = True
    if registros_existentes > 0:
        st.error(f"⚠️ ¡Alto ahí! Ya existen {registros_existentes} registros para la {input_semana} en la base de datos. Sincronización bloqueada para evitar duplicados.")
        puede_subir = False
    else:
        st.success(f"✅ La {input_semana} está libre en Supabase. Lista para recibir datos.")
    
    st.divider()
    
    col3, col4 = st.columns(2)
    with col3:
        base_mod3 = st.file_uploader("Sube la base 'AMTC (6)'", type=['xlsx'], key="mod3_base")
    with col4:
        reporte_mod3 = st.file_uploader("Sube el Reporte Southbound", type=['xlsx'], key="mod3_target")

    if base_mod3 and reporte_mod3 and puede_subir:
        with st.spinner('Procesando datos y validando reglas de negocio...'):
            # --- Lógica de diccionarios y cruce (Igual a la versión anterior) ---
            df_base = pd.read_excel(base_mod3, sheet_name='TJ-LB')
            base_dict = {}
            for index, row in df_base.iterrows():
                container = str(row.iloc[0]).strip()
                if container == 'nan' or not container: continue
                base_dict[container] = {
                    'gate_out': clean_date(row.get('GATE OUT LB')),
                    'empty_term': clean_date(row.get('Empty Termination (EI)')),
                    'amtc_delivery': clean_val(row.get('AMTC Delivery (Loaded)')),
                    'amtc_unload': clean_val(row.get('AMTC Unload Notification (Empty)'))
                }
            
            df = pd.read_excel(reporte_mod3)
            if 'CONTAINER' not in [str(c).upper() for c in df.columns]:
                df = pd.read_excel(reporte_mod3, header=4)
                
            df.columns = [str(c).strip().upper() for c in df.columns]
            df['CONTAINER'] = df['CONTAINER'].astype(str).str.strip()
            df = df[df['CONTAINER'].str.len() == 11]
            df = df.map(clean_val)
            
            return_col_name = 'TERMINAL RETURN EMPTY' if 'TERMINAL RETURN EMPTY' in df.columns else 'RETURN EMPTY DATE'
            if return_col_name not in df.columns and 'TERMINAL EMPTY RETURN VALIDATION' in df.columns:
                 return_col_name = 'TERMINAL EMPTY RETURN VALIDATION'
                 
            column_mapping = {
                'CONTAINER': 'CONTAINER',
                'PICK UP DATE': 'PICK UP DATE',
                'AMTC Delivery (Loaded)': 'MTC MX DELIVERY',
                'AMTC Unload Notification (Empty)': 'MTC EMPTY NOTIFICATION',
                return_col_name: 'TERMINAL EMPTY RETURN VALIDATION ',
                'CHASIS DAYS INBOUND': 'CHASIS DAYS INBOUND',
                'TOTAL': 'TOTAL'
            }
            
            df_sql = df.rename(columns=column_mapping)
            df_sql['SEMANA'] = input_semana
            
            st.markdown("#### Vista previa de la Ingesta:")
            cols_to_show = [c for c in df_sql.columns if c in column_mapping.values() or c == 'SEMANA']
            st.dataframe(df_sql[cols_to_show].head())
            
            if st.button(f"☁️ Sincronizar {input_semana} con Supabase", type="primary"):
                with st.spinner("Escribiendo en PostgreSQL..."):
                    df_upload = df_sql[cols_to_show].copy()
                    for col in df_upload.select_dtypes(include=['datetime64', 'datetimetz']).columns:
                         df_upload[col] = df_upload[col].dt.strftime('%Y-%m-%d %H:%M:%S')
                         
                    records = df_upload.where(pd.notnull(df_upload), None).to_dict(orient='records')
                    try:
                        supabase.table('mtc_southbound').insert(records).execute()
                        st.success(f"¡Ingesta completada! {len(records)} registros guardados.")
                        st.balloons()
                    except Exception as e:
                        st.error(f"Fallo en la escritura a la base de datos: {e}")
